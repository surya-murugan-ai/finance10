"""
Document processing service for file upload and validation
"""
import os
import pandas as pd
from typing import Dict, Any, Optional, List
from fastapi import UploadFile, HTTPException
import magic
from io import BytesIO
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from ..config import settings

class DocumentProcessor:
    """Handle document processing and validation"""
    
    ALLOWED_EXTENSIONS = {'.pdf', '.xlsx', '.xls', '.csv'}
    ALLOWED_MIME_TYPES = {
        'application/pdf',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
        'text/csv'
    }
    
    def __init__(self):
        # Create uploads directory if it doesn't exist
        os.makedirs(settings.UPLOAD_FOLDER, exist_ok=True)
    
    async def process_file(self, file: UploadFile, user_id: str) -> Dict[str, Any]:
        """Process uploaded file and extract data"""
        
        # Validate file
        await self._validate_file(file)
        
        # Read file content
        content = await file.read()
        
        # Determine file type and extract data
        if file.filename.endswith('.pdf'):
            data = await self._process_pdf(content)
        elif file.filename.endswith(('.xlsx', '.xls')):
            data = await self._process_excel(content)
        elif file.filename.endswith('.csv'):
            data = await self._process_csv(content)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type")
        
        # Infer document type
        document_type = self._infer_document_type(file.filename)
        
        return {
            "type": document_type,
            "size": len(content),
            "data": data,
            "filename": file.filename
        }
    
    async def _validate_file(self, file: UploadFile) -> None:
        """Validate uploaded file"""
        if not file.filename:
            raise HTTPException(status_code=400, detail="No filename provided")
        
        # Check file extension
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in self.ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"File type not allowed. Allowed types: {', '.join(self.ALLOWED_EXTENSIONS)}"
            )
        
        # Check file size
        file.file.seek(0, 2)  # Seek to end
        file_size = file.file.tell()
        file.file.seek(0)  # Reset to beginning
        
        if file_size > settings.MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"File too large. Maximum size: {settings.MAX_FILE_SIZE / (1024*1024):.1f}MB"
            )
    
    async def _process_pdf(self, content: bytes) -> Dict[str, Any]:
        """Process PDF file"""
        try:
            pdf_reader = PdfReader(BytesIO(content))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            
            return {
                "text": text,
                "page_count": len(pdf_reader.pages),
                "type": "pdf"
            }
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error processing PDF: {str(e)}")
    
    async def _process_excel(self, content: bytes) -> Dict[str, Any]:
        """Process Excel file"""
        try:
            # Read Excel file
            workbook = load_workbook(BytesIO(content))
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Convert to DataFrame
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                
                if data:
                    # Create DataFrame
                    df = pd.DataFrame(data[1:], columns=data[0])
                    sheets_data[sheet_name] = df.to_dict('records')
            
            return {
                "sheets": sheets_data,
                "sheet_count": len(sheets_data),
                "type": "excel"
            }
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error processing Excel: {str(e)}")
    
    async def _process_csv(self, content: bytes) -> Dict[str, Any]:
        """Process CSV file"""
        try:
            # Read CSV file
            df = pd.read_csv(BytesIO(content))
            
            return {
                "data": df.to_dict('records'),
                "columns": df.columns.tolist(),
                "row_count": len(df),
                "type": "csv"
            }
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error processing CSV: {str(e)}")
    
    def _infer_document_type(self, filename: str) -> str:
        """Infer document type from filename"""
        name = filename.lower()
        
        if any(keyword in name for keyword in ['vendor', 'invoice', 'bill']):
            return 'vendor_invoice'
        elif any(keyword in name for keyword in ['sales', 'register', 'receipt']):
            return 'sales_register'
        elif any(keyword in name for keyword in ['salary', 'payroll', 'wage']):
            return 'salary_register'
        elif any(keyword in name for keyword in ['bank', 'statement', 'passbook']):
            return 'bank_statement'
        elif any(keyword in name for keyword in ['purchase', 'procurement', 'po']):
            return 'purchase_register'
        elif any(keyword in name for keyword in ['journal', 'entry', 'jv']):
            return 'journal_entry'
        elif any(keyword in name for keyword in ['gst', 'tax', 'return']):
            return 'gst_return'
        elif any(keyword in name for keyword in ['tds', 'deduction', 'withholding']):
            return 'tds_certificate'
        else:
            return 'general_document'
    
    def get_sample_data(self, document_type: str, filename: str) -> Dict[str, Any]:
        """Generate sample data for demonstration"""
        sample_data = {
            'vendor_invoice': {
                'invoices': [
                    {
                        'invoice_number': 'VI-2025-001',
                        'vendor_name': 'TechCorp Solutions',
                        'invoice_date': '2025-01-15',
                        'amount': 125000.00,
                        'gstin': '09ABCDE1234F1Z5',
                        'status': 'pending'
                    }
                ]
            },
            'sales_register': {
                'sales': [
                    {
                        'sale_id': 'S-2025-001',
                        'customer_name': 'Global Enterprises',
                        'sale_date': '2025-01-16',
                        'amount': 250000.00,
                        'tax_amount': 45000.00,
                        'total_amount': 295000.00
                    }
                ]
            },
            'salary_register': {
                'employees': [
                    {
                        'employee_id': 'EMP001',
                        'employee_name': 'John Doe',
                        'department': 'Engineering',
                        'basic_salary': 75000.00,
                        'tds_deducted': 7500.00,
                        'net_salary': 67500.00
                    }
                ]
            }
        }
        
        return sample_data.get(document_type, {'message': 'Document processed successfully'})
    
    def get_supported_formats(self) -> List[str]:
        """Get supported file formats"""
        return list(self.ALLOWED_EXTENSIONS)